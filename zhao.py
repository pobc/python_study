"""
程序思路
# 1.安装、导入工具包
pip3 install openpyxl
# 2.打开excel
# 3.阅读内容
# 4.找到包含abc的记录
# 5.做统计
# 6.输出结果
"""


# 1.安装、导入工具包
import openpyxl

# 1.1 找到文件路径
wenjian_lujin = r"C:\Users\Jiang\Desktop\data.xlsx"
# 1.2 要查找的关键字
guanjianzi="abc"
# 2.打开excel
xm_excel = openpyxl.load_workbook(wenjian_lujin)
# 3.阅读内容 ,内容在Sheet1中
xm_sheet1 = xm_excel['Sheet1']
# 我们从0开始数
tongji = 0
# 4.找到包含abc的记录
for i in range(xm_sheet1.max_row):
    hang1 = xm_sheet1.cell(row=i + 1, column=1).value
    if guanjianzi in hang1:
        # 5.做统计
        tongji = tongji + 1

# 6.输出结果
print("abc的统计结果为：")
print(tongji)